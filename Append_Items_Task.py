from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, Border, Alignment
from copy import copy

wb = load_workbook('test2.xlsx')
ws = wb['Sheet1']

#initiating indexes
cols=0
row=0
i=0
row_data = []
sheetNames = ['Sheet2', 'Sheet3', 'Sheet4'] #to name the new sheets created
data_range = 'C5:L93' #specify the range of the data in Excel
item_array = ['ITEM1', 'ITEM2', 'ITEM3']
num_rows = 93 #number of rows in dataset


#we want to iterate over the entire dataset in Sheet1
for row in range(0,num_rows-5): # num_rows-5 because dataset starts from row 5
    for cols in range(0,10):
        #if the value of any cell is the same as the string 'ITEM1' run the following
        if ws[data_range][row][cols].value==item_array[0]:
            ws2 = wb.create_sheet(sheetNames[i]) #create new sheet upon finding 'ITEM1'
            ws2 = wb[sheetNames[i]]
            print("New Sheet with label " +sheetNames[i]+ " created")
            i+=1
            #to copy the data, we must re-align the data to its original cell
            for row_2 in range(row,row+11):
                for cols_2 in range(cols,cols+10):
                    #print("Exporting data " + "("+str(row)+","+str(cols)+")" + " from Sheet1 to " + "("+str(row_2)+","+str(cols_2)+")" + " in " + sheetNames[i-1])
                    ws2[data_range][row_2-row][cols_2-cols].value = ws[data_range][row_2][cols_2].value
                    #copy styling
                    ws2[data_range][row_2-row][cols_2-cols].font = copy(ws[data_range][row_2][cols_2].font)
                    ws2[data_range][row_2-row][cols_2-cols].fill = copy(ws[data_range][row_2][cols_2].fill)
                    ws2[data_range][row_2-row][cols_2-cols].border = copy(ws[data_range][row_2][cols_2].border)
                    ws2[data_range][row_2-row][cols_2-cols].number_format = copy(ws[data_range][row_2][cols_2].number_format)
            #if the string 'ITEM2' is found, then we want to add this with the previous data of 'ITEM1' (with the same date)

        #loop to add more items to the dataset of 'ITEM1', then append them to new sheet - loops through item_array to add item data
        shift=0
        for j in range(1, len(item_array)):
            if ws[data_range][row][cols].value==item_array[j]:
                for row_2 in range(row,row+11):
                    for cols_2 in range(cols,cols+10):
                        ws2['C19:L45'][row_2-row+shift][cols_2-cols].value = ws[data_range][row_2][cols_2].value
                        #copy styling
                        ws2['C19:L45'][row_2-row+shift][cols_2-cols].font = copy(ws[data_range][row_2][cols_2].font)
                        ws2['C19:L45'][row_2-row+shift][cols_2-cols].fill = copy(ws[data_range][row_2][cols_2].fill)
                        ws2['C19:L45'][row_2-row+shift][cols_2-cols].border = copy(ws[data_range][row_2][cols_2].border)
                        ws2['C19:L45'][row_2-row+shift][cols_2-cols].number_format = copy(ws[data_range][row_2][cols_2].number_format)
            shift+=14 #to space out the data and prevent overwrite

#save it into a new file
wb.save('test4.xlsx')
